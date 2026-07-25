"""
excel_io.py
Excel export/import, working with in-memory byte streams so FastAPI can
stream a download or accept an upload without touching local disk.
"""
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

import storage

COLUMNS = ["ID", "Title", "Priority", "Status", "Created At"]


def export_to_excel_bytes(status: str = "all") -> bytes:
    tasks = storage.list_tasks(status=status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for t in tasks:
        ws.append([t["id"], t["title"], t["priority"], t["status"], t["created_at"]])

    widths = [6, 45, 12, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_from_excel_bytes(file_bytes: bytes, skip_duplicates: bool = True):
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    def col_index(*names):
        for name in names:
            if name in header:
                return header.index(name)
        return None

    title_idx = col_index("title", "task", "name")
    priority_idx = col_index("priority")
    status_idx = col_index("status")

    if title_idx is None:
        raise ValueError("Could not find a 'Title' column in the Excel file.")

    existing_titles = {t["title"].lower() for t in storage.list_tasks()} if skip_duplicates else set()

    imported = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[title_idx] if title_idx < len(row) else None
        if not title or not str(title).strip():
            continue
        title = str(title).strip()

        if skip_duplicates and title.lower() in existing_titles:
            skipped += 1
            continue

        priority = "medium"
        if priority_idx is not None and priority_idx < len(row) and row[priority_idx]:
            p = str(row[priority_idx]).strip().lower()
            if p in ("low", "medium", "high"):
                priority = p

        task_id = storage.add_task(title, priority)

        if status_idx is not None and status_idx < len(row) and row[status_idx]:
            s = str(row[status_idx]).strip().lower()
            if s in ("done", "completed", "complete"):
                storage.complete_task(task_id)

        existing_titles.add(title.lower())
        imported += 1

    return imported, skipped
